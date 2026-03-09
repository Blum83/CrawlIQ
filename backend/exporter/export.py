"""
Export QA report to HTML, Excel, or CSV.
"""

import csv
import io
from datetime import datetime


# ─── HTML ─────────────────────────────────────────────────────────────────────

def export_html(url: str, result: dict) -> str:
    issues = result.get("issues", {})
    cov = result.get("content_coverage", {})
    pages = result.get("page_details", [])
    ai_summary = result.get("ai_summary", "")
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    def cnt(key):
        return issues.get(key, {}).get("count", 0)

    def badge(val, warn=1, crit=5):
        if val == 0:
            return f'<span class="badge ok">{val}</span>'
        if val < crit:
            return f'<span class="badge warn">{val}</span>'
        return f'<span class="badge crit">{val}</span>'

    def issue_rows(key, sub="urls"):
        items = issues.get(key, {}).get(sub, [])
        if not items:
            return "<tr><td colspan='2' class='none'>None</td></tr>"
        rows = []
        for item in items:
            if isinstance(item, dict):
                rows.append(f"<tr><td>{item.get('url','')}</td><td>{item.get('status','')}</td></tr>")
            else:
                rows.append(f"<tr><td colspan='2'>{item}</td></tr>")
        return "\n".join(rows)

    pages_crawled = result.get("pages_crawled", 0)
    total_issues = sum(cnt(k) for k in [
        "missing_title", "missing_meta_description", "missing_h1", "multiple_h1",
        "missing_canonical", "duplicate_titles", "missing_html_lang",
        "missing_alt_tags", "buttons_missing_label", "broken_images",
        "thin_content_under_200_words", "empty_pages", "non_200_status"
    ])

    score = max(0, 100 - int(total_issues / max(pages_crawled, 1) * 100))
    if score >= 80:   health_cls, health_txt = "ok", f"Good ({score}/100)"
    elif score >= 50: health_cls, health_txt = "warn", f"Needs work ({score}/100)"
    else:             health_cls, health_txt = "crit", f"Poor ({score}/100)"

    page_rows = ""
    for p in pages:
        sc = p.get("status_code", 0)
        sc_cls = "" if sc == 200 else ("warn" if sc in (301, 302) else "crit")
        page_rows += f"""<tr>
            <td><a href="{p['url']}" target="_blank">{p['url']}</a></td>
            <td class="{sc_cls}">{sc}</td>
            <td>{"✅" if p.get("has_title") else "❌"}</td>
            <td>{"✅" if p.get("has_meta_description") else "❌"}</td>
            <td>{"✅" if p.get("has_h1") else "❌"}</td>
            <td>{"✅" if p.get("has_canonical") else "❌"}</td>
            <td>{p.get("images_missing_alt", 0)}</td>
            <td>{p.get("word_count", 0)}</td>
        </tr>"""

    ai_block = ""
    if ai_summary:
        ai_block = f"""
        <section>
            <h2>🤖 AI Summary</h2>
            <div class="ai-summary">{ai_summary.replace(chr(10), "<br>")}</div>
        </section>"""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>QA Report — {url}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #0f1117; color: #e2e8f0; padding: 2rem; }}
  h1 {{ font-size: 1.6rem; margin-bottom: 0.25rem; }}
  h2 {{ font-size: 1.1rem; margin: 1.5rem 0 0.75rem; color: #a78bfa; }}
  a {{ color: #818cf8; }}
  .meta {{ color: #64748b; font-size: 0.85rem; margin-bottom: 2rem; }}
  section {{ background: #1e2130; border-radius: 10px; padding: 1.25rem 1.5rem; margin-bottom: 1.25rem; }}
  .overview-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 1rem; margin-top: 0.75rem; }}
  .stat {{ background: #252a3d; border-radius: 8px; padding: 1rem; text-align: center; }}
  .stat .value {{ font-size: 1.8rem; font-weight: 700; }}
  .stat .label {{ font-size: 0.75rem; color: #94a3b8; margin-top: 0.25rem; }}
  .health.ok {{ color: #4ade80; }} .health.warn {{ color: #facc15; }} .health.crit {{ color: #f87171; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; margin-top: 0.5rem; }}
  th {{ background: #252a3d; padding: 0.5rem 0.75rem; text-align: left; color: #94a3b8; font-weight: 600; }}
  td {{ padding: 0.45rem 0.75rem; border-bottom: 1px solid #2a3050; word-break: break-all; }}
  tr:last-child td {{ border-bottom: none; }}
  .badge {{ display: inline-block; padding: 0.15rem 0.5rem; border-radius: 999px; font-weight: 600; font-size: 0.8rem; }}
  .badge.ok {{ background: #14532d; color: #4ade80; }}
  .badge.warn {{ background: #713f12; color: #facc15; }}
  .badge.crit {{ background: #7f1d1d; color: #f87171; }}
  .issues-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(300px, 1fr)); gap: 1rem; }}
  .issue-card {{ background: #252a3d; border-radius: 8px; padding: 1rem; }}
  .issue-card h3 {{ font-size: 0.85rem; color: #94a3b8; margin-bottom: 0.5rem; }}
  td.ok {{ color: #4ade80; }} td.warn {{ color: #facc15; }} td.crit {{ color: #f87171; }}
  td.none {{ color: #475569; font-style: italic; }}
  .ai-summary {{ background: #252a3d; border-radius: 8px; padding: 1rem; line-height: 1.7; color: #cbd5e1; }}
  .footer {{ text-align: center; color: #334155; font-size: 0.8rem; margin-top: 2rem; }}
</style>
</head>
<body>
<h1>🔍 QA Report</h1>
<p class="meta">🌐 <a href="{url}" target="_blank">{url}</a> &nbsp;·&nbsp; Generated {generated}</p>

<section>
  <h2>📊 Overview</h2>
  <div class="overview-grid">
    <div class="stat"><div class="value">{pages_crawled}</div><div class="label">Pages Crawled</div></div>
    <div class="stat"><div class="value">{result.get('error_pages', 0)}</div><div class="label">Error Pages</div></div>
    <div class="stat"><div class="value">{cov.get('avg_word_count', 0)}</div><div class="label">Avg Words/Page</div></div>
    <div class="stat"><div class="value">{total_issues}</div><div class="label">Total Issues</div></div>
    <div class="stat"><div class="value health {health_cls}">{health_txt}</div><div class="label">Health Score</div></div>
  </div>
</section>

<section>
  <h2>🔎 SEO</h2>
  <table>
    <tr><th>Check</th><th>Count</th></tr>
    <tr><td>Missing title</td><td>{badge(cnt("missing_title"))}</td></tr>
    <tr><td>Missing meta description</td><td>{badge(cnt("missing_meta_description"))}</td></tr>
    <tr><td>Missing H1</td><td>{badge(cnt("missing_h1"))}</td></tr>
    <tr><td>Multiple H1</td><td>{badge(cnt("multiple_h1"))}</td></tr>
    <tr><td>Missing canonical</td><td>{badge(cnt("missing_canonical"), 5, 20)}</td></tr>
    <tr><td>Duplicate titles</td><td>{badge(cnt("duplicate_titles"))}</td></tr>
  </table>
</section>

<section>
  <h2>♿ Accessibility</h2>
  <table>
    <tr><th>Check</th><th>Count</th></tr>
    <tr><td>Missing HTML lang</td><td>{badge(cnt("missing_html_lang"), 1, 3)}</td></tr>
    <tr><td>Images without alt text</td><td>{badge(cnt("missing_alt_tags"), 1, 10)}</td></tr>
    <tr><td>Buttons without label</td><td>{badge(cnt("buttons_missing_label"))}</td></tr>
  </table>
</section>

<section>
  <h2>📄 Content &amp; Technical</h2>
  <table>
    <tr><th>Check</th><th>Count</th></tr>
    <tr><td>Thin content (&lt;200 words)</td><td>{badge(cnt("thin_content_under_200_words"), 3, 10)}</td></tr>
    <tr><td>Empty pages</td><td>{badge(cnt("empty_pages"))}</td></tr>
    <tr><td>Broken images</td><td>{badge(cnt("broken_images"), 1, 10)}</td></tr>
    <tr><td>Non-200 status pages</td><td>{badge(cnt("non_200_status"))}</td></tr>
  </table>
</section>

<section>
  <h2>🔗 Issue Details</h2>
  <div class="issues-grid">
    <div class="issue-card"><h3>Pages without meta description</h3><table>{issue_rows("missing_meta_description")}</table></div>
    <div class="issue-card"><h3>Pages without H1</h3><table>{issue_rows("missing_h1")}</table></div>
    <div class="issue-card"><h3>Pages with multiple H1</h3><table>{issue_rows("multiple_h1")}</table></div>
    <div class="issue-card"><h3>Duplicate titles</h3><table>{issue_rows("duplicate_titles")}</table></div>
    <div class="issue-card"><h3>Non-200 pages</h3><table><tr><th>URL</th><th>Status</th></tr>{issue_rows("non_200_status", "pages")}</table></div>
  </div>
</section>

{ai_block}

<section>
  <h2>📋 All Pages</h2>
  <table>
    <tr><th>URL</th><th>Status</th><th>Title</th><th>Meta</th><th>H1</th><th>Canonical</th><th>Missing Alt</th><th>Words</th></tr>
    {page_rows}
  </table>
</section>

<div class="footer">AI QA Agent · {generated}</div>
</body>
</html>"""


# ─── Excel ────────────────────────────────────────────────────────────────────

def export_excel(url: str, result: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    issues = result.get("issues", {})

    def cnt(key):
        return issues.get(key, {}).get("count", 0)

    # ── helpers ──
    HDR_FILL = PatternFill("solid", fgColor="1E2130")
    HDR_FONT = Font(bold=True, color="A78BFA")
    OK_FILL  = PatternFill("solid", fgColor="14532D")
    WRN_FILL = PatternFill("solid", fgColor="713F12")
    CRT_FILL = PatternFill("solid", fgColor="7F1D1D")
    OK_FONT  = Font(color="4ADE80")
    WRN_FONT = Font(color="FACC15")
    CRT_FONT = Font(color="F87171")

    def fill_font(val, warn=1, crit=5):
        if val == 0:      return OK_FILL, OK_FONT
        if val < crit:    return WRN_FILL, WRN_FONT
        return CRT_FILL, CRT_FONT

    def header_row(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center")

    def autowidth(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 80)

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = "Overview"
    header_row(ws, ["Metric", "Value"])
    rows = [
        ("Website", url),
        ("Pages Crawled", result.get("pages_crawled", 0)),
        ("Error Pages", result.get("error_pages", 0)),
        ("Avg Word Count", result.get("content_coverage", {}).get("avg_word_count", 0)),
        ("Thin Content %", result.get("content_coverage", {}).get("pct_thin_content", 0)),
        ("", ""),
        ("— SEO —", ""),
        ("Missing Title", cnt("missing_title")),
        ("Missing Meta Description", cnt("missing_meta_description")),
        ("Missing H1", cnt("missing_h1")),
        ("Multiple H1", cnt("multiple_h1")),
        ("Missing Canonical", cnt("missing_canonical")),
        ("Duplicate Titles", cnt("duplicate_titles")),
        ("", ""),
        ("— Accessibility —", ""),
        ("Missing HTML Lang", cnt("missing_html_lang")),
        ("Images Without Alt", cnt("missing_alt_tags")),
        ("Buttons Without Label", cnt("buttons_missing_label")),
        ("", ""),
        ("— Content & Technical —", ""),
        ("Thin Content Pages", cnt("thin_content_under_200_words")),
        ("Empty Pages", cnt("empty_pages")),
        ("Broken Images", cnt("broken_images")),
        ("Non-200 Pages", cnt("non_200_status")),
    ]
    for label, val in rows:
        ws.append([label, val])
        if isinstance(val, int) and val >= 0 and label.startswith(("Missing", "Multiple", "Broken", "Thin", "Empty", "Buttons", "Duplicate")):
            f, fn = fill_font(val)
            ws.cell(ws.max_row, 2).fill = f
            ws.cell(ws.max_row, 2).font = fn
    autowidth(ws)

    # ── Sheet 2: SEO Issues ──
    ws2 = wb.create_sheet("SEO Issues")
    header_row(ws2, ["Category", "URL"])
    for key, label in [
        ("missing_title", "Missing Title"),
        ("missing_meta_description", "Missing Meta"),
        ("missing_h1", "Missing H1"),
        ("multiple_h1", "Multiple H1"),
        ("missing_canonical", "Missing Canonical"),
        ("duplicate_titles", "Duplicate Title"),
    ]:
        for u in issues.get(key, {}).get("urls", []):
            ws2.append([label, u])
    autowidth(ws2)

    # ── Sheet 3: Accessibility ──
    ws3 = wb.create_sheet("Accessibility")
    header_row(ws3, ["Category", "URL / Detail"])
    for key, label in [
        ("missing_html_lang", "Missing Lang"),
        ("missing_alt_tags", "Images w/o Alt"),
        ("buttons_missing_label", "Button w/o Label"),
    ]:
        sub = "urls" if key == "missing_html_lang" else "pages"
        for u in issues.get(key, {}).get(sub, []):
            ws3.append([label, u])
    autowidth(ws3)

    # ── Sheet 4: Technical ──
    ws4 = wb.create_sheet("Technical")
    header_row(ws4, ["Category", "URL", "Status Code"])
    for p in issues.get("non_200_status", {}).get("pages", []):
        ws4.append(["Non-200", p.get("url", ""), p.get("status", "")])
    for u in issues.get("broken_images", {}).get("pages", []):
        ws4.append(["Broken Image", u, ""])
    autowidth(ws4)

    # ── Sheet 5: All Pages ──
    ws5 = wb.create_sheet("All Pages")
    header_row(ws5, ["URL", "Status", "Has Title", "Has Meta", "Has H1", "Has Canonical", "Missing Alt", "Words", "Error"])
    for p in result.get("page_details", []):
        ws5.append([
            p.get("url", ""),
            p.get("status_code", 0),
            "Yes" if p.get("has_title") else "No",
            "Yes" if p.get("has_meta_description") else "No",
            "Yes" if p.get("has_h1") else "No",
            "Yes" if p.get("has_canonical") else "No",
            p.get("images_missing_alt", 0),
            p.get("word_count", 0),
            p.get("error", "") or "",
        ])
    autowidth(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── CSV ──────────────────────────────────────────────────────────────────────

def export_csv(result: dict) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["URL", "Status Code", "Has Title", "Has Meta Description", "Has H1",
                "Has Canonical", "H1 Count", "Images Missing Alt", "Word Count", "Error"])
    for p in result.get("page_details", []):
        w.writerow([
            p.get("url", ""),
            p.get("status_code", ""),
            p.get("has_title", ""),
            p.get("has_meta_description", ""),
            p.get("has_h1", ""),
            p.get("has_canonical", ""),
            p.get("h1_count", ""),
            p.get("images_missing_alt", ""),
            p.get("word_count", ""),
            p.get("error", "") or "",
        ])
    return out.getvalue()
